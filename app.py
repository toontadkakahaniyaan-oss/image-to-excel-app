"""
DocuMorph AI
============
Converts any uploaded image (tables, receipts, invoices, handwritten notes,
or flowcharts) into a perfectly structured, editable Excel sheet using a
multimodal Vision LLM (via the OpenAI SDK, pointed at OpenAI or OpenRouter).

Run with:
    streamlit run app.py
"""

import base64
import io
import json
import re
import time
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image
from openai import OpenAI


# =========================================================================
# PAGE CONFIG
# =========================================================================

st.set_page_config(
    page_title="DocuMorph AI",
    page_icon="🪄",
    layout="wide",
    initial_sidebar_state="expanded",
)


# =========================================================================
# CUSTOM CSS — premium dark, card-based UI
# =========================================================================

CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=Space+Grotesk:wght@600;700&display=swap');

html, body, [class*="css"], .stMarkdown, p, span, label, div {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
}

.stApp {
    background: radial-gradient(circle at 15% 0%, #1b1f3b 0%, #0c0e1a 45%, #060710 100%);
}

#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header[data-testid="stHeader"] {background: transparent;}

/* ---------- Hero ---------- */
.dm-hero {
    background: linear-gradient(135deg, rgba(99,102,241,0.18) 0%, rgba(168,85,247,0.10) 50%, rgba(236,72,153,0.08) 100%);
    border: 1px solid rgba(148,163,184,0.14);
    border-radius: 22px;
    padding: 30px 36px;
    margin-bottom: 26px;
    box-shadow: 0 8px 32px rgba(0,0,0,0.35);
}
.dm-hero-title {
    font-family: 'Space Grotesk', sans-serif;
    font-size: 2.15rem;
    font-weight: 700;
    color: #f8fafc;
    margin: 0;
    letter-spacing: -0.02em;
}
.dm-hero-title span {
    background: linear-gradient(90deg, #818cf8, #e879f9, #f472b6);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}
.dm-hero-subtitle {
    color: #94a3b8;
    font-size: 0.98rem;
    margin-top: 8px;
    max-width: 680px;
    line-height: 1.5;
}
.dm-hero-pills { margin-top: 14px; }
.dm-hero-pills span {
    background: rgba(255,255,255,0.06);
    border: 1px solid rgba(148,163,184,0.18);
    padding: 4px 13px;
    border-radius: 999px;
    font-size: 0.78rem;
    color: #cbd5e1;
    margin-right: 8px;
    display: inline-block;
    margin-top: 8px;
}

/* ---------- Card containers (st.container(border=True)) ---------- */
div[data-testid="stVerticalBlockBorderWrapper"] {
    background: linear-gradient(180deg, rgba(30,34,58,0.55), rgba(18,20,36,0.55));
    border: 1px solid rgba(148,163,184,0.14) !important;
    border-radius: 18px !important;
    backdrop-filter: blur(6px);
}
div[data-testid="stVerticalBlockBorderWrapper"] > div { padding: 4px 2px; }

/* ---------- Sidebar ---------- */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0c0e1c 0%, #090a14 100%);
    border-right: 1px solid rgba(148,163,184,0.1);
}
section[data-testid="stSidebar"] h3, section[data-testid="stSidebar"] h4 {
    color: #e2e8f0 !important;
    font-family: 'Space Grotesk', sans-serif;
}

/* ---------- Buttons ---------- */
.stButton > button, .stDownloadButton > button {
    border-radius: 12px;
    font-weight: 600;
    border: 1px solid rgba(148,163,184,0.18);
    transition: all 0.15s ease;
}
button[kind="primary"] {
    background: linear-gradient(90deg, #6366f1, #a855f7) !important;
    border: none !important;
    box-shadow: 0 4px 18px rgba(99,102,241,0.35);
    color: #fff !important;
}
button[kind="primary"]:hover {
    box-shadow: 0 6px 24px rgba(168,85,247,0.5);
    transform: translateY(-1px);
}
.stDownloadButton > button {
    background: rgba(30,34,58,0.85);
    color: #e2e8f0;
}
.stDownloadButton > button:hover {
    border-color: #818cf8;
    color: #fff;
}

/* ---------- Inputs ---------- */
.stTextInput input, .stTextArea textarea, div[data-baseweb="select"] > div {
    background-color: rgba(15,17,32,0.7) !important;
    border-radius: 10px !important;
    border: 1px solid rgba(148,163,184,0.2) !important;
    color: #e2e8f0 !important;
}
.stTextInput input:focus, .stTextArea textarea:focus {
    border-color: #818cf8 !important;
    box-shadow: 0 0 0 2px rgba(129,140,248,0.25) !important;
}

/* ---------- Metrics ---------- */
div[data-testid="stMetric"] {
    background: rgba(99,102,241,0.08);
    border: 1px solid rgba(99,102,241,0.18);
    border-radius: 14px;
    padding: 10px 14px 6px 14px;
}
div[data-testid="stMetricValue"] { color: #f8fafc; font-family: 'Space Grotesk', sans-serif; }
div[data-testid="stMetricLabel"] { color: #94a3b8; }

/* ---------- File uploader ---------- */
section[data-testid="stFileUploaderDropzone"] {
    background: rgba(99,102,241,0.05);
    border: 1.5px dashed rgba(129,140,248,0.4);
    border-radius: 14px;
}

/* ---------- Data editor / dataframe ---------- */
div[data-testid="stDataFrame"], div[data-testid="stDataEditor"] {
    border-radius: 14px;
    overflow: hidden;
    border: 1px solid rgba(148,163,184,0.15);
}

/* ---------- Expander ---------- */
details {
    background: rgba(15,17,32,0.5);
    border-radius: 12px !important;
    border: 1px solid rgba(148,163,184,0.12) !important;
}

/* ---------- Scrollbar ---------- */
::-webkit-scrollbar { width: 10px; height: 10px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: rgba(129,140,248,0.35); border-radius: 10px; }
::-webkit-scrollbar-thumb:hover { background: rgba(129,140,248,0.55); }

h4, h5 { font-family: 'Space Grotesk', sans-serif; color: #f1f5f9 !important; font-weight: 600; }

.dm-footer {
    text-align: center;
    color: #64748b;
    font-size: 0.82rem;
    margin-top: 34px;
    padding-bottom: 10px;
}
</style>
"""

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


# =========================================================================
# CONSTANTS — modes & prompts
# =========================================================================

MODE_PROMPTS: Dict[str, Dict[str, str]] = {
    "table": {
        "label": "🗂️ Structured Table (Auto-detect)",
        "instruction": (
            "You are analyzing an image of a table. This may be a financial statement, "
            "invoice, receipt, spreadsheet screenshot, printed report, or any tabular data.\n"
            "Extract EVERY row and column exactly as shown, preserving the original reading order.\n"
            "Rules:\n"
            "- Preserve exact header names as written in the image.\n"
            "- If a cell visually spans multiple columns or rows (a merged cell), repeat its value "
            "in every cell it covers so no cell is left blank purely because of merging.\n"
            "- Preserve numeric values exactly as printed (do not round, do not add or remove currency "
            "symbols, keep thousands separators as shown).\n"
            "- Do not invent, summarize, skip, or reorder any row.\n"
            "- If a cell is genuinely empty in the image, use an empty string for it.\n"
        ),
    },
    "flowchart": {
        "label": "🔀 Flowchart / Process Sequence",
        "instruction": (
            "You are analyzing an image of a flowchart, process diagram, swimlane diagram, or "
            "decision tree.\n"
            "Convert every visual node, arrow, and branch into a sequential table with EXACTLY "
            "these column headers, in this exact order:\n"
            'Step No | Process / Action | Condition / Decision | Next Step | Responsible / Notes\n'
            "Rules:\n"
            "- Number steps in the logical order they would actually be followed "
            "(top-to-bottom, left-to-right, following the arrows).\n"
            "- For decision diamonds, put the question/condition text in 'Condition / Decision' and "
            "describe both outgoing branches in 'Next Step' "
            "(e.g. 'If Yes -> Step 4, If No -> Step 6').\n"
            "- If a swimlane, role label, or department indicates who owns a step, capture it in "
            "'Responsible / Notes'; otherwise leave that cell empty.\n"
            "- Do not skip any node, arrow label, start/end terminator, or branch.\n"
        ),
    },
    "keyvalue": {
        "label": "🔑 Key-Value / Form Extractor",
        "instruction": (
            "You are analyzing an image of a form, application, ID document, invoice header, or "
            "any label-and-value layout.\n"
            "Extract every field into EXACTLY two columns with these headers:\n"
            "Field Name | Value\n"
            "Rules:\n"
            "- One row per field. Use the field's printed label as 'Field Name'.\n"
            "- If a field has no value filled in, use an empty string for 'Value'.\n"
            "- Preserve values exactly as written (dates, numbers, IDs). For checkboxes/tick marks, "
            "use 'Checked' or 'Unchecked'.\n"
            "- Do not merge unrelated fields into a single row, and do not skip any visible field.\n"
        ),
    },
    "custom": {
        "label": "✍️ Custom Prompt / Instruction Mode",
        "instruction": (
            "You are analyzing the provided image and extracting structured data from it "
            "according to the user's custom instructions below. Still return the result as a "
            "clean, well-organized table.\n"
        ),
    },
}

JSON_FORMAT_INSTRUCTION = """
CRITICAL OUTPUT FORMAT RULES — follow exactly:
Respond with ONLY a single valid JSON object and absolutely nothing else — no explanations, no
markdown code fences, no commentary before or after it.

The JSON object MUST have exactly this shape:
{
  "headers": ["Column 1", "Column 2", "..."],
  "rows": [
    ["row1 col1 value", "row1 col2 value", "..."],
    ["row2 col1 value", "row2 col2 value", "..."]
  ]
}

- Every entry in "rows" MUST be an array with the same number of elements as "headers".
- All values must be plain strings (convert numbers to strings, e.g. "1,234.50").
- This JSON array format exists specifically so commas inside cell text (e.g. addresses, notes,
  "Rice, 5kg") never get confused with column separators the way they would in raw CSV text.
- Do not wrap the JSON in ```json fences and do not add any prose. Output raw JSON only.
"""

SYSTEM_PROMPT = (
    "You are DocuMorph AI's vision extraction engine. You convert images of tables, forms, "
    "receipts, handwritten notes, and flowcharts into precise, structured JSON data for Excel "
    "export. You are meticulous about not missing rows or fields, and you never add commentary — "
    "you output only valid JSON as instructed."
)


# =========================================================================
# HELPER FUNCTIONS — image prep
# =========================================================================

def prepare_image(uploaded_file) -> Tuple[Image.Image, str]:
    """Load, normalize, and base64-encode the uploaded image for the Vision API."""
    uploaded_file.seek(0)
    image = Image.open(uploaded_file)
    image.load()
    if image.mode in ("RGBA", "P", "LA"):
        image = image.convert("RGB")
    elif image.mode not in ("RGB", "L"):
        image = image.convert("RGB")

    max_dim = 2200
    if max(image.size) > max_dim:
        ratio = max_dim / max(image.size)
        new_size = (max(1, int(image.size[0] * ratio)), max(1, int(image.size[1] * ratio)))
        image = image.resize(new_size, Image.Resampling.LANCZOS)

    buffer = io.BytesIO()
    image.save(buffer, format="PNG", optimize=True)
    b64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return image, b64


# =========================================================================
# HELPER FUNCTIONS — Vision API call
# =========================================================================

def call_vision_model(
    api_key: str,
    base_url: str,
    model: str,
    user_prompt: str,
    image_b64: str,
    extra_headers: Optional[Dict[str, str]] = None,
    temperature: float = 0.1,
    max_tokens: int = 4096,
) -> str:
    client = OpenAI(api_key=api_key, base_url=base_url, default_headers=extra_headers or {}, timeout=90.0)

    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {
            "role": "user",
            "content": [
                {"type": "text", "text": user_prompt},
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:image/png;base64,{image_b64}"},
                },
            ],
        },
    ]

    response = client.chat.completions.create(
        model=model,
        messages=messages,
        max_tokens=max_tokens,
        temperature=temperature,
    )
    return response.choices[0].message.content or ""


def friendly_error(e: Exception) -> str:
    msg = str(e)
    lower = msg.lower()
    if "401" in msg or "invalid api key" in lower or "unauthorized" in lower or "authentication" in lower:
        return "Invalid API key. Please double-check the key in the sidebar."
    if "429" in msg or "rate limit" in lower or "quota" in lower:
        return (
            "Rate limit or quota reached. If you're on a free OpenRouter model, wait a minute "
            "(free tier ≈ 20 requests/min, 50/day) or try again shortly."
        )
    if "timeout" in lower or "timed out" in lower:
        return "The request timed out. Try again, use a smaller/clearer image, or switch models."
    if "connection" in lower or "network" in lower:
        return "Couldn't connect to the API. Check your internet connection or the provider/base URL."
    if "cannot identify image" in lower or "unidentifiedimage" in lower:
        return "The uploaded file doesn't look like a valid image. Please try a different file."
    if "model" in lower and ("not found" in lower or "does not exist" in lower or "invalid" in lower):
        return "That model ID wasn't recognized by the provider. Please check the model name in the sidebar."
    return f"{msg}"


# =========================================================================
# HELPER FUNCTIONS — robust response parsing (JSON -> Markdown table fallback)
# =========================================================================

def clean_json_text(text: str) -> str:
    text = text.strip()
    fence_match = re.search(r"```(?:json)?\s*([\s\S]*?)```", text)
    if fence_match:
        return fence_match.group(1).strip()
    return text


def extract_json_snippet(text: str) -> Optional[str]:
    start_candidates = [i for i in [text.find("{"), text.find("[")] if i != -1]
    if not start_candidates:
        return None
    start = min(start_candidates)
    end_curly = text.rfind("}")
    end_square = text.rfind("]")
    end = max(end_curly, end_square)
    if end == -1 or end < start:
        return None
    return text[start:end + 1]


def dedupe_columns(cols: List[Any]) -> List[str]:
    seen: Dict[str, int] = {}
    result = []
    for c in cols:
        c = str(c).strip() if c not in (None, "") else "Column"
        if c == "":
            c = "Column"
        if c in seen:
            seen[c] += 1
            result.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 0
            result.append(c)
    return result


def json_to_dataframe(parsed: Any) -> Optional[pd.DataFrame]:
    """Handles both the requested {'headers','rows'} shape and common model deviations
    (plain list-of-objects, plain 2D array)."""
    try:
        if isinstance(parsed, dict) and "headers" in parsed and "rows" in parsed:
            headers = dedupe_columns([str(h) for h in parsed["headers"]])
            rows = parsed["rows"]
            norm_rows = []
            for r in rows:
                if isinstance(r, dict):
                    r = [r.get(h, "") for h in headers]
                else:
                    r = list(r)
                if len(r) < len(headers):
                    r = r + [""] * (len(headers) - len(r))
                elif len(r) > len(headers):
                    r = r[: len(headers)]
                norm_rows.append(["" if c is None else str(c) for c in r])
            if not norm_rows:
                return pd.DataFrame(columns=headers)
            return pd.DataFrame(norm_rows, columns=headers)

        elif isinstance(parsed, list) and len(parsed) > 0 and isinstance(parsed[0], dict):
            df = pd.DataFrame(parsed)
            df = df.fillna("").astype(str)
            df.columns = dedupe_columns(list(df.columns))
            return df

        elif isinstance(parsed, list) and len(parsed) > 0 and isinstance(parsed[0], list):
            header = parsed[0]
            data = parsed[1:]
            max_cols = max(len(header), max((len(r) for r in data), default=0))
            header = list(header) + [f"Column {i + 1}" for i in range(len(header), max_cols)]
            header = dedupe_columns(header)
            norm_rows = []
            for r in data:
                r = list(r)
                if len(r) < max_cols:
                    r = r + [""] * (max_cols - len(r))
                norm_rows.append(["" if c is None else str(c) for c in r[:max_cols]])
            if not norm_rows:
                return pd.DataFrame(columns=header)
            return pd.DataFrame(norm_rows, columns=header)
    except Exception:
        return None
    return None


def markdown_table_to_df(text: str) -> Optional[pd.DataFrame]:
    lines = [l for l in text.splitlines() if l.strip().startswith("|")]
    if len(lines) < 2:
        return None
    sep_pattern = re.compile(r"^\|?\s*:?-{2,}:?\s*(\|\s*:?-{2,}:?\s*)+\|?$")
    rows = []
    for line in lines:
        if sep_pattern.match(line.strip()):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        rows.append(cells)
    if len(rows) < 1:
        return None
    max_cols = max(len(r) for r in rows)
    rows = [r + [""] * (max_cols - len(r)) for r in rows]
    header, *data = rows
    header = dedupe_columns(header)
    if not data:
        return pd.DataFrame(columns=header)
    return pd.DataFrame(data, columns=header)


def parse_ai_response(raw_text: str) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
    """Returns (dataframe_or_None, warning_message_or_None). Never raises."""
    if not raw_text or not raw_text.strip():
        return None, "The model returned an empty response."

    cleaned = clean_json_text(raw_text)

    parsed = None
    try:
        parsed = json.loads(cleaned)
    except (json.JSONDecodeError, ValueError):
        snippet = extract_json_snippet(cleaned)
        if snippet:
            try:
                parsed = json.loads(snippet)
            except (json.JSONDecodeError, ValueError):
                parsed = None

    if parsed is not None:
        df = json_to_dataframe(parsed)
        if df is not None:
            return df, None

    md_df = markdown_table_to_df(raw_text)
    if md_df is not None:
        return md_df, (
            "The model didn't return clean JSON, so DocuMorph AI recovered the data from a "
            "Markdown table instead. Please double-check the results below."
        )

    return None, (
        "Couldn't parse a structured table from the model's response. See the raw output below "
        "and try again — perhaps with a different mode, a clearer image, or another model."
    )


# =========================================================================
# HELPER FUNCTIONS — export
# =========================================================================

def dataframe_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "DocuMorph Export") -> bytes:
    output = io.BytesIO()
    safe_sheet_name = (sheet_name or "Sheet1")[:31]
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=safe_sheet_name)
        worksheet = writer.sheets[safe_sheet_name]

        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
        thin_side = Side(style="thin", color="D1D5DB")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        for col_cells in worksheet.columns:
            max_len = 0
            for c in col_cells:
                if c.value is not None:
                    max_len = max(max_len, len(str(c.value)))
            col_letter = col_cells[0].column_letter
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

        if worksheet.max_row >= 1:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

    return output.getvalue()


def dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8-sig")


def clipboard_button(df: pd.DataFrame) -> None:
    tsv = df.to_csv(index=False, sep="\t")
    tsv_js = json.dumps(tsv)
    html_code = f"""
    <div style="font-family:'Inter',sans-serif;">
      <button id="dm-copy-btn" onclick="dmCopyTable()" style="
          width:100%; padding:9px 14px; border-radius:12px;
          background: rgba(30,34,58,0.9); color:#e2e8f0;
          border:1px solid rgba(148,163,184,0.35); font-weight:600;
          font-size:14px; cursor:pointer; transition: all .15s ease;">
        📋 Copy Table
      </button>
    </div>
    <script>
      function dmCopyTable() {{
        const text = {tsv_js};
        const btn = document.getElementById('dm-copy-btn');
        navigator.clipboard.writeText(text).then(() => {{
          btn.innerText = '✅ Copied!';
          btn.style.borderColor = '#818cf8';
          setTimeout(() => {{ btn.innerText = '📋 Copy Table'; }}, 1800);
        }}).catch(() => {{
          btn.innerText = '⚠️ Copy failed — select manually';
          setTimeout(() => {{ btn.innerText = '📋 Copy Table'; }}, 2200);
        }});
      }}
    </script>
    """
    components.html(html_code, height=48)


# =========================================================================
# SESSION STATE
# =========================================================================

_DEFAULTS = {
    "extracted_df": None,
    "raw_response": "",
    "warning_msg": None,
    "processing_time": None,
    "last_mode_label": None,
    "converted_once": False,
}
for _k, _v in _DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v


def reset_results():
    st.session_state.extracted_df = None
    st.session_state.raw_response = ""
    st.session_state.warning_msg = None
    st.session_state.processing_time = None
    st.session_state.last_mode_label = None
    st.session_state.converted_once = False


# =========================================================================
# SIDEBAR — engine configuration
# =========================================================================

with st.sidebar:
    st.markdown("### ⚙️ Engine Configuration")

    provider = st.selectbox("Provider", ["OpenRouter", "OpenAI"], index=0)
    api_key = st.text_input(
        "API Key",
        type="password",
        placeholder="sk-...",
        help="Used only for this session's requests — never stored or logged.",
    )

    if provider == "OpenRouter":
        base_url = "https://openrouter.ai/api/v1"
        model_options = [
            "nvidia/nemotron-nano-12b-v2-vl:free  🆓",
            "openrouter/free  🆓 (auto-router)",
            "google/gemini-2.5-flash",
            "openai/gpt-4o",
            "anthropic/claude-3.5-sonnet",
            "qwen/qwen2.5-vl-72b-instruct",
            "Custom...",
        ]
    else:
        base_url = "https://api.openai.com/v1"
        model_options = ["gpt-4o", "gpt-4o-mini", "gpt-4.1", "Custom..."]

    model_choice = st.selectbox("Vision Model", model_options, index=0)
    if model_choice == "Custom...":
        model = st.text_input("Custom model ID", placeholder="e.g. mistralai/pixtral-large-2411")
    else:
        # strip the display-only "🆓" badge / notes to get the real model ID sent to the API
        model = model_choice.split("  ")[0].strip()

    if provider == "OpenRouter" and "🆓" in model_choice:
        st.success(
            "This model is free on OpenRouter — no credit card needed. "
            "Free tier is rate-limited (~20 req/min, 50/day), so give it a few seconds between conversions."
        )

    with st.expander("🎛️ Advanced settings"):
        temperature = st.slider("Temperature", 0.0, 1.0, 0.1, 0.05, help="Lower = more literal/consistent extraction.")
        max_tokens = st.slider("Max output tokens", 1024, 8192, 4096, 512)

    st.markdown("---")
    if provider == "OpenRouter":
        st.caption(
            "🆓 Get a free key (no card) at [openrouter.ai/keys](https://openrouter.ai/keys) — "
            "sign in with email or GitHub, generate a key, paste it above."
        )
    st.caption("🔒 DocuMorph AI never stores your images or API keys — everything runs within this browser session.")

    if st.session_state.converted_once:
        st.markdown("---")
        if st.button("🔄 Clear results", use_container_width=True):
            reset_results()
            st.rerun()


# =========================================================================
# HERO HEADER
# =========================================================================

st.markdown(
    """
    <div class="dm-hero">
        <div class="dm-hero-title">🪄 DocuMorph <span>AI</span></div>
        <div class="dm-hero-subtitle">
            Turn any image — tables, receipts, invoices, handwritten notes, or flowcharts —
            into a perfectly structured, editable Excel sheet in seconds.
        </div>
        <div class="dm-hero-pills">
            <span>🗂️ Tables</span><span>🧾 Receipts</span><span>🔀 Flowcharts</span>
            <span>🔑 Forms</span><span>✍️ Handwritten Notes</span>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)


# =========================================================================
# MAIN LAYOUT — split screen
# =========================================================================

left_col, right_col = st.columns([1, 1.35], gap="large")

# ---------------- LEFT: Upload & Mode Selection ----------------
with left_col:
    with st.container(border=True):
        st.markdown("#### 📤 Upload & Configure")

        uploaded_file = st.file_uploader(
            "Upload an image",
            type=["png", "jpg", "jpeg", "webp", "bmp"],
            help="Tables, receipts, invoices, handwritten notes, or flowchart images work best.",
        )

        if uploaded_file is not None:
            st.image(uploaded_file, use_container_width=True, caption=uploaded_file.name)

        mode_key = st.selectbox(
            "Conversion Mode",
            options=list(MODE_PROMPTS.keys()),
            format_func=lambda k: MODE_PROMPTS[k]["label"],
        )

        custom_instruction = ""
        if mode_key == "custom":
            custom_instruction = st.text_area(
                "Custom Instructions",
                placeholder="e.g. Group items by category and add a totals row at the bottom...",
                height=110,
            )

        convert_clicked = st.button(
            "✨ Convert to Structured Data",
            type="primary",
            use_container_width=True,
            disabled=uploaded_file is None,
        )

    if convert_clicked and uploaded_file is not None:
        if not api_key:
            st.error("⚠️ Please add your API key in the sidebar before converting.")
        elif not model:
            st.error("⚠️ Please select or enter a model in the sidebar.")
        else:
            with st.spinner("🔍 DocuMorph AI is reading your image and extracting structured data..."):
                try:
                    start_time = time.time()
                    _, image_b64 = prepare_image(uploaded_file)

                    mode_info = MODE_PROMPTS[mode_key]
                    prompt = mode_info["instruction"]
                    if mode_key == "custom":
                        if custom_instruction.strip():
                            prompt += f"\nUser's custom instructions:\n{custom_instruction.strip()}\n"
                        else:
                            prompt += "\nExtract all visible structured data as a clean table.\n"
                    prompt += JSON_FORMAT_INSTRUCTION

                    extra_headers = {}
                    if provider == "OpenRouter":
                        extra_headers = {
                            "HTTP-Referer": "https://documorph.ai",
                            "X-Title": "DocuMorph AI",
                        }

                    raw_text = call_vision_model(
                        api_key=api_key,
                        base_url=base_url,
                        model=model,
                        user_prompt=prompt,
                        image_b64=image_b64,
                        extra_headers=extra_headers,
                        temperature=temperature,
                        max_tokens=max_tokens,
                    )
                    elapsed = time.time() - start_time

                    df, warning = parse_ai_response(raw_text)

                    st.session_state.raw_response = raw_text
                    st.session_state.warning_msg = warning
                    st.session_state.processing_time = elapsed
                    st.session_state.last_mode_label = mode_info["label"]
                    st.session_state.converted_once = True

                    if df is not None:
                        st.session_state.extracted_df = df
                        if warning:
                            st.toast("Recovered data via fallback parser", icon="⚠️")
                        else:
                            st.toast("Extraction complete!", icon="✅")
                    else:
                        st.session_state.extracted_df = pd.DataFrame({"Column 1": [""]})
                        st.toast("Couldn't parse a clean table — check raw output", icon="⚠️")

                except Exception as e:
                    st.session_state.converted_once = True
                    st.session_state.extracted_df = None
                    st.session_state.raw_response = ""
                    st.session_state.processing_time = None
                    st.session_state.warning_msg = f"❌ Extraction failed: {friendly_error(e)}"

            st.rerun()


# ---------------- RIGHT: Extracted Table & Export ----------------
with right_col:
    with st.container(border=True):
        st.markdown("#### 📊 Extracted Data")

        if st.session_state.extracted_df is not None:
            df = st.session_state.extracted_df

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Rows", len(df))
            m2.metric("Columns", len(df.columns))
            m3.metric(
                "Time",
                f"{st.session_state.processing_time:.1f}s" if st.session_state.processing_time else "—",
            )
            mode_display = (
                st.session_state.last_mode_label.split(" ", 1)[-1]
                if st.session_state.last_mode_label
                else "—"
            )
            m4.metric("Mode", mode_display)

            if st.session_state.warning_msg:
                st.warning(st.session_state.warning_msg)

            edited_df = st.data_editor(
                df,
                use_container_width=True,
                num_rows="dynamic",
                key="dm_data_editor",
                height=420,
            )
            st.session_state.extracted_df = edited_df

            st.markdown("##### 📥 Export")
            e1, e2, e3 = st.columns(3)
            with e1:
                st.download_button(
                    "⬇️ Excel (.xlsx)",
                    data=dataframe_to_excel_bytes(edited_df),
                    file_name="documorph_export.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with e2:
                st.download_button(
                    "⬇️ CSV",
                    data=dataframe_to_csv_bytes(edited_df),
                    file_name="documorph_export.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
            with e3:
                clipboard_button(edited_df)

            with st.expander("🔎 View raw model response"):
                st.code(st.session_state.raw_response or "—", language="json")

        elif st.session_state.converted_once:
            st.error(st.session_state.warning_msg or "Extraction failed. Please try again.")
            if st.session_state.raw_response:
                with st.expander("🔎 View raw model response"):
                    st.code(st.session_state.raw_response, language="text")

        else:
            st.info("👈 Upload an image, pick a conversion mode, and click **Convert** to see your structured table here.")


# =========================================================================
# FOOTER
# =========================================================================

st.markdown(
    '<div class="dm-footer">Built with ❤️ using Streamlit &amp; Vision AI — DocuMorph AI</div>',
    unsafe_allow_html=True,
)
